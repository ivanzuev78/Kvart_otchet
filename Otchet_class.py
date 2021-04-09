import datetime
import itertools
import os
from copy import deepcopy

import openpyxl as opx
from openpyxl.styles import Alignment, Border, Font, Side


class Worker:
    """
    Объект Работник.
    Хранит в себе всё, что сделал.
    """

    def __init__(self, name):
        # Переменные для всего, что работник сделал
        self.name = name
        self.plenki = []
        self.obrazci = []
        self.otcheti = []
        self.naneseniya = []
        self.sinthesis = []


class Shablon:
    # Класс-родитель для объектов работы (плёнки, жидкие образцы, отчёты, синтезы)
    def __init__(self, date, author, tema, markirovka=None):
        # Данные об объекте
        self.tema = tema
        self.date = date
        self.author = author
        self.markirovka = markirovka
        self.old_tema = None

    def __str__(self):
        return self.markirovka

    def __repr__(self):
        return self.markirovka

    def __iter__(self):
        for i in self.markirovka:
            yield i


class Plenka(Shablon):
    def __init__(self, date, author, tema, markirovka):
        super().__init__(date, author, tema, markirovka)


class Obrazec(Shablon):
    def __init__(self, date, author, tema, markirovka):
        super().__init__(date, author, tema, markirovka)


class Otchet(Shablon):
    def __init__(self, date, author, tema, markirovka):
        super().__init__(date, author, tema, markirovka)


class Nanesenie(Shablon):
    def __init__(self, date, author, tema, markirovka):
        super().__init__(date, author, tema, markirovka)


class Sintez(Shablon):
    def __init__(self, date, author, tema, komponent, mass):
        super().__init__(date, author, tema)
        if isinstance(mass, (int, float)):
            self.mass = mass
        else:
            try:
                self.mass = float(mass)
                if self.mass < 0:
                    raise ValueError
            except ValueError:
                self.mass = 0
        self.komponent = komponent
        self.markirovka = tema + komponent + str(self.mass)
        self.base_tema = tema


class Cunductor:
    """
    Класс - обработчик всей информации со всех файлов
    """

    def __init__(
        self,
        date_start=None,
        date_end=None,
        svodnaya_name_file="Не указано",
        production_name_file="Не указано",
        report_name_file="Не указано",
        sintez_name_file="Не указано",
    ):

        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        self._date_start = date_start  # Дата начала квартала
        self._date_end = date_end  # Дата окончания квартала
        self.svodnaya_name_file = svodnaya_name_file  # Путь к сводной таблице
        self.production_name_file = (
            production_name_file  # Путь к Общему перечню продукции ОВНТ
        )
        self.report_name_file = report_name_file  # Путь к общему списку отчётов
        self.sintez_name_file = sintez_name_file  # Пусть к списку синтезов

        self.thems = {}  # Название темы: Объект класса Tema
        self.all_thems = []  # Список всех тем, в выбранном квартале
        self.thems_replased = {}  # Название переименнованной темы: Объект класса Tema

        self.global_tems = (
            {}
        )  # Название глобальной темы: Список тем, которые на неё заменятся

        # Данные файлов Excel в виде массива массивов
        self.svod_tabl_ws = []
        self.report_ws = []
        self.production_ws = {}  # Название листа: массив данных
        self.sintez_ws = []

        self.ignor_tema = []  # Список всех когда-либо добавленных в игнор тем
        self.current_ignor_tema = (
            []
        )  # Список добавленных в игнор тем, которые присутствуют в данном квартале
        self.text_of_tema = (
            {}
        )  # Название темы: Текст, который подставится в отчёте с именами вместо этой темы
        self.text_of_tema_noname = (
            {}
        )  # Название темы: Текст, который подставится в отчёте без имён вместо этой темы
        self.names = []  # Имена всех сотрудников, кто работал в текущем квартале
        self.ignor_names = []  # Имена для игнорирования

        self.all_data = []  # Список всех данных. Объекты класса Shablon

        # Список имен для замены в Общем перечне отчётов ОВНТ
        self.good_names = {}

    @property
    def date_start(self):
        return self._date_start

    @property
    def date_end(self):
        return self._date_end

    def change_date(self, start, end):
        """
        Функция для изменения дат в кондукторе
        :param start: [Год, месяц, день]
        :param end: [Год, месяц, день]
        :return: None
        """
        import datetime

        self._date_start = datetime.datetime(start[0], start[1], start[2])
        self._date_end = datetime.datetime(end[0], end[1], end[2])

    @staticmethod
    def check_date_type(date):
        """
        Функция проверяет, является ли клетка датой
        :param date: datetime.datetime
        :return: bool

        """

        if isinstance(date, datetime.datetime):
            return True
        elif isinstance(date, str):
            try:
                datetime.datetime(int(date[6:]), int(date[3:5]), int(date[:2]))
                return True
            except:
                return False
        else:
            return False

    def check_date(self, date):
        """
        Функция проверяет, принадлежит ли дата кварталу
        :param date: datetime.datetime
        :return: bool
        """
        if isinstance(date, str):
            try:
                date = datetime.datetime(int(date[6:]), int(date[3:5]), int(date[:2]))
            except:
                return False
        if isinstance(date, datetime.datetime):
            if self._date_start <= date <= self._date_end:
                return True
            else:
                return False
        else:
            return False

    def input_name(self, name: str) -> str:
        """
        :param name: str (Фамилия И.О. в любом формате)
        :return: str (Фамилия И.О. с пробелом после фамилии)
        """

        if name in self.good_names:
            name = self.good_names[name]

        integ = ""  # Возвращаемая строка
        probel_check = False  # Флаг для постановки пробела после фамилии
        prev = ""  # Предыдущий символ при проходе ФИО
        for index, i in enumerate(name):  # Проходим по всем элементам
            if i == "." and not probel_check:  # Если встречаем точку
                integ += " "  # Вставляем пробел перед добавлением предыдущего символа
                probel_check = True  # Меняем флаг
            if i != " ":  # Если символ не пробел
                integ += prev  # Добавляем предыдущий
                prev = i  # Запоминаем текущий
        integ += prev  # Добавляем последний в конце
        return integ  # Возвращаем ФИО

    # Обработка данных из сводной таблицы
    def svod_tabl_count(
        self,
        tabl,
        markirovka_col=0,
        date_col=4,
        author_col=5,
        tema_col=1,
        type_naneseniya_col=3,
    ):
        """
                Обрабатывает данные из сводной таблицы и добавляет их в self.all_data в качестве объекта класса Shablon
        0) Маркировка плёнки
        1) Тема
        2) Номер темы
        3) Тип нанесения
        4) Дата
        5) Ответственный

        :param tabl:
        :param markirovka_col:
        :param date_col:
        :param author_col:
        :param tema_col:
        :param type_naneseniya_col:
        :return:
        """
        for index, row in enumerate(tabl):  # Берем строку
            if type(row[tema_col]) is not str:
                row[tema_col] = str(row[tema_col])
            if self.check_date_type(row[date_col]):
                if row[author_col]:
                    row[author_col] = self.input_name(
                        row[author_col]
                    )  # Форматируем ФИО по шаблону

                if (
                    isinstance(row[type_naneseniya_col], str)
                    and "аборатор" in row[type_naneseniya_col]
                ):
                    self.all_data.append(
                        Plenka(
                            row[date_col],
                            row[author_col],
                            row[tema_col],
                            row[markirovka_col],
                        )
                    )
                elif (
                    isinstance(row[type_naneseniya_col], str)
                    and "ромыш" in row[type_naneseniya_col]
                ):
                    self.all_data.append(
                        Nanesenie(
                            row[date_col],
                            row[author_col],
                            row[tema_col],
                            row[markirovka_col],
                        )
                    )

    # Обработка данных из перечня отчётов
    def report_count(
        self, tabl, date_col=5, author_col=2, nomer_otcheta_col=0, tema_col=6
    ):
        """
        Обрабатывает данные из списка отчётов и добавляет их в self.all_data в качестве объекта класса Shablon
            0) Номер отчёта
            1) Название отчёта
            2) Ответственный 1
            3) Ответственный 2
            4) Ответственный 3
            5) Дата
            6) Название темы
            :param tabl:
            :return:
        """
        for row in tabl:  # Берем строку
            if type(row[tema_col]) is not str:
                row[tema_col] = str(row[tema_col])
            if self.check_date_type(row[date_col]):
                row[author_col] = self.input_name(
                    row[author_col]
                )  # Форматируем ФИО по шаблону
                if row[nomer_otcheta_col]:  # Если есть номер отчёта
                    self.all_data.append(
                        Otchet(
                            row[date_col],
                            row[author_col],
                            row[tema_col],
                            row[nomer_otcheta_col],
                        )
                    )

    # Обработка данных из общего перечня продукции ОВНТ
    def production_count(self, tabl: opx.Workbook.active, name: str) -> None:
        """
        Обрабатывает данные из общего перечня продукции ОВНТ
        и добавляет их в self.all_data в качестве объекта класса Shablon

        0) Маркировка состава
        1) Дата
        2) Тема
        :param tabl: Полученная таблица
        :param name: Имя сотрудника
        :return: None
        """

        col_flag = [False, False, False]
        mark_col = 0
        date_col = 1
        tema_col = 2

        for row in tabl:  # Берем строку
            if not all(col_flag):
                for col, cell in enumerate(row):
                    if cell:
                        if type(cell) is str and "аркиров" in cell:
                            mark_col = col
                            col_flag[0] = True
                        if type(cell) is str and ("Дата" in cell or "дата" in cell):
                            date_col = col
                            col_flag[1] = True
                        if type(cell) is str and ("истем" in cell or "рецепт" in cell):
                            tema_col = col
                            col_flag[2] = True
            if type(row[tema_col]) is not str and all(col_flag):
                row[tema_col] = str(row[tema_col])
            if self.check_date_type(row[date_col]):
                if row[tema_col]:  # Если клетка с темой не пустая
                    self.all_data.append(
                        Obrazec(row[date_col], name, row[tema_col], row[mark_col])
                    )

    def sintez_count(
        self, tabl, tema_col=0, sostav_col=2, date_col=5, mass_col=6, author_col=8
    ):
        """
        Обрабатывает данные из списка синтезов и добавляет их в self.all_data в качестве объекта класса Shablon
        0) Тема
        1) Номер темы
        2) Состав (основа / отвердитель)
        3) Рецептура
        4) Номер техкарты
        5) Дата
        6) Масса
        7) Партия
        8) Ответственный
        :param tabl:
        :return:
        """

        for row in tabl:  # Берем строку
            if type(row[tema_col]) is not str:
                row[tema_col] = str(row[tema_col])
            if self.check_date_type(row[date_col]):
                if row[author_col]:
                    row[author_col] = self.input_name(
                        row[author_col]
                    )  # Форматируем ФИО по шаблону
                self.all_data.append(
                    Sintez(
                        row[date_col],
                        row[author_col],
                        row[tema_col],
                        row[sostav_col],
                        row[mass_col],
                    )
                )

    # Считывает данные страницы Excel
    @staticmethod
    def read_excel(ws, col: int = 10):
        """
        Функция для чтения excel файла
        :param ws: Лист страницы Excel, с которого надо получить данные
        :param col: количество считаных столбцов
        :return: Список массивов с элементами ячеек
        """
        sv_tabl = []  # Список массивов, который вернем

        for row in ws:  # Проходим по всем строкам на листе
            current_row = []  # Память для текущего ряда

            # Проходим по всем ячейкам в строке. Нумеруем, что бы не считывать все колонки
            for index, cell in enumerate(row):
                if index == col:
                    break
                current_row.append(cell.value)  # Добавляем значение ячейки в массив
            if any(current_row):
                sv_tabl.append(current_row)  # Добавляем массив в список
        return sv_tabl

    # Читает фалы Excel
    def reader(self):
        """
        Читает файли Excel и сохраняет их в виде массивов для дальнейшей обработки
        :return: None
        """

        yield 10, "Сводная таблица"
        svod_wb = opx.load_workbook(filename=self.svodnaya_name_file)
        svod_ws = svod_wb["Сводная таблица"]
        self.svod_tabl_ws = self.read_excel(svod_ws)

        yield 20, "Отчёты"
        report_wb = opx.load_workbook(filename=self.report_name_file)
        report_ws = report_wb["Отчеты"]
        self.report_ws = self.read_excel(report_ws, col=7)

        yield 40, "Общий перечень продукции ОВНТ"
        production_wb = opx.load_workbook(filename=self.production_name_file)
        prod_names = production_wb.sheetnames
        process = 40
        for name in prod_names:
            process += 5
            self.production_ws[name] = self.read_excel(production_wb[name], col=5)
            yield process, f"Общий перечень продукции ОВНТ: {name}"

        yield 80, "Список синтезов"
        sintez_wb = opx.load_workbook(filename=self.sintez_name_file)
        sintez_ws = sintez_wb["Данные"]
        self.sintez_ws = self.read_excel(sintez_ws, col=9)

        yield 100, "Готово!"

    def counter(self):
        """
        Обрабатывает данные распарсеных таблиц
        :return: None
        """
        self.all_data = []
        self.svod_tabl_count(self.svod_tabl_ws)
        self.report_count(self.report_ws)
        for name in self.production_ws:
            self.production_count(self.production_ws[name], self.input_name(name))
        self.sintez_count(self.sintez_ws)

    def sort_data(self):
        """
        Смотрит кучу файлов и распределяет их по темам
        :return: None
        """
        self.thems = {}
        self.all_thems = []
        self.current_ignor_tema = []
        self.names = []
        for itemx in self.all_data:
            item = deepcopy(itemx)
            if not self.check_date(item.date):
                continue
            if item.tema not in self.all_thems:
                self.all_thems.append(item.tema)
            if item.author not in self.names:
                self.names.append(item.author)
            if item.tema in self.ignor_tema:
                if item.tema not in self.current_ignor_tema:
                    self.current_ignor_tema.append(item.tema)
                continue
            elif item.author in self.ignor_names:
                continue
            if item.tema in self.thems_replased:
                item.tema = self.thems_replased[item.tema]
            if item.tema not in self.thems:
                self.thems[item.tema] = Tema(item.tema)
            self.thems[item.tema].add_item(item)

    @staticmethod
    def short_show(massiv):
        """
        Функция позволяет сворачивать штучные образцы в дефисные группы
        :param massiv: Массив с образцами или плёнками поштучно
        :return: Массив с образцами или плёнками, свёрнуто дефисными группами
        """
        fine = []
        current_prefix = ""
        current_numb = -1
        first_numb = -1
        sorted_massiv = sorted(massiv, key=lambda x: x.markirovka)
        for word in sorted_massiv:
            word_prefix = ""
            word_numb = ""
            for b in word.markirovka:
                if b.isdigit():
                    word_numb += b
                else:
                    word_prefix += b
            if (
                word_prefix == current_prefix
                and int(current_numb) == int(word_numb) - 1
            ):
                current_numb = word_numb
                if word == sorted_massiv[-1]:
                    if current_numb != first_numb:
                        fine.append(f"{current_prefix}{first_numb}-{current_numb}")
                    else:
                        fine.append(f"{current_prefix}{first_numb}")
            else:
                if current_prefix:
                    if current_numb != first_numb:
                        fine.append(f"{current_prefix}{first_numb}-{current_numb}")
                    else:
                        fine.append(f"{current_prefix}{first_numb}")
                current_prefix = word_prefix
                first_numb = word_numb
                current_numb = first_numb
                if word == sorted_massiv[-1] and current_prefix:
                    if current_numb != first_numb:
                        fine.append(f"{current_prefix}{first_numb}-{current_numb}")
                    else:
                        fine.append(f"{current_prefix}{first_numb}")
        return fine

    @staticmethod
    def short_show_report(massiv):
        """
        Функция, которая оставляет только первые цифры отчёта.
        Отсекает на первой не цифре
        :param massiv: Массив с отчётами
        :return: Массив с номерами отчётов
        """
        massiv_to_return = []
        for otchet in massiv:
            word = ""
            for b in otchet:
                if b.isdigit():
                    word += b
                else:
                    break
            massiv_to_return.append(word)

        return massiv_to_return

    @staticmethod
    def okonchanie(word, numb):
        """
        Функция для правильного написания окончания к числу
        :param word: Слово для правильного склонения
        :param numb: Число, к которому нужно подобрать окончание
        :return: Число + слово в нужном падеже
        """
        dict_word = {
            "образец": {
                "Род.падеж.множ.число": "образцов",
                "Род.падеж.ед.число": "образца",
                "Имен.падеж.ед.число": "образец",
            },
            "нанесение": {
                "Род.падеж.множ.число": "нанесений",
                "Род.падеж.ед.число": "нанесения",
                "Имен.падеж.ед.число": "нанесение",
            },
            "отчёт": {
                "Род.падеж.множ.число": "отчётов",
                "Род.падеж.ед.число": "отчёта",
                "Имен.падеж.ед.число": "отчёт",
            },
            "синтез": {
                "Род.падеж.множ.число": "синтезов",
                "Род.падеж.ед.число": "синтеза",
                "Имен.падеж.ед.число": "синтез",
            },
        }

        if numb % 100 in [i for i in range(11, 15)] or numb % 10 in [0, 5, 6, 7, 8, 9]:
            return f"{numb} " + dict_word[word]["Род.падеж.множ.число"]
        if numb % 10 in [2, 3, 4]:
            return f"{numb} " + dict_word[word]["Род.падеж.ед.число"]
        else:
            return f"{numb} " + dict_word[word]["Имен.падеж.ед.число"]

    def make_excel(self, kvartal, year, personal_date=False):
        """
        Функция, которая создает отчёт с именами.
        :param personal_date: bool - особое поведение функции, если особые даты
        :param kvartal: Номер квартала (справочно, для подставления в название отчёта)
        :param year: Номер года (справочно, для подставления в название отчёта)
        :return: Путь до файла
        """
        wb = opx.Workbook()
        ws_title = [
            "Выпущенные опытно-промышленные партии",
            "Изготовленные лабораторные образцы и компоненты",
            "Написанные отчеты",
            "Выпущенные рецептуры и ТК",
            "Проведенные промышленные нанесения",
        ]
        ws = wb.active
        ws_count_row = []

        for tema in self.thems:
            if tema in self.text_of_tema:
                ws_title.insert(0, self.text_of_tema[tema])
            else:
                ws_title.insert(0, tema)
            ws.append(ws_title)
            ws_title.pop(0)
            ws_count_row.append(True)
            total_obraz = 0
            total_otchet = 0
            total_nanesenie = 0
            total_sintes_mass = 0
            total_sintes_count = 0
            for name in self.thems[tema].workers:
                sum_obraz = len(self.thems[tema].workers[name].plenki) + len(
                    self.thems[tema].workers[name].obrazci
                )
                total_obraz += sum_obraz
                sum_otchet = len(self.thems[tema].workers[name].otcheti)
                total_otchet += sum_otchet
                sum_nanesenie = len(self.thems[tema].workers[name].naneseniya)
                total_nanesenie += sum_nanesenie
                current_list_to_append = [name]
                sintez_counter = SintezCounter()
                for i in self.thems[tema].workers[name].sinthesis:
                    sintez_counter.add_sintez(i)
                total_sintes_mass += sintez_counter.return_all()[0]
                total_sintes_count += sintez_counter.return_all()[1]
                cell = sintez_counter.str_one_tema()
                current_list_to_append.append(cell)
                cell = ""
                for i in self.short_show(
                    self.thems[tema].workers[name].plenki
                ) + self.short_show(self.thems[tema].workers[name].obrazci):
                    cell += f"{i}, "
                if cell:
                    cell = cell[:-2]
                    cell += f"\n\nИтого: " + self.okonchanie("образец", sum_obraz)
                current_list_to_append.append(cell)  # Добавление образцов и плёнок
                cell = ""
                for i in self.short_show_report(self.thems[tema].workers[name].otcheti):
                    cell += f"{i}, "
                if cell:
                    cell = cell[:-2]
                    cell += f"\n\nИтого: " + self.okonchanie("отчёт", sum_otchet)
                current_list_to_append.append(cell)  # Добавление отчётов

                current_list_to_append.append(
                    " "
                )  # Добавление Выпущенные рецептуры и ТК

                cell = ""
                for i in self.short_show(self.thems[tema].workers[name].naneseniya):
                    cell += f"{i}, "
                if cell:
                    cell = cell[:-2]
                    cell += f"\n\nИтого: " + self.okonchanie("нанесение", sum_nanesenie)
                current_list_to_append.append(cell)  # Добавление нанесения
                ws.append(current_list_to_append)
                ws_count_row.append(True)
            total_string = ["ИТОГО"]
            if total_sintes_count:
                total_string.append(
                    self.okonchanie("синтез", total_sintes_count)
                    + f", {total_sintes_mass} кг"
                )
            else:
                total_string.append("")
            if total_obraz:
                total_string.append(self.okonchanie("образец", total_obraz))
            else:
                total_string.append("")
            if total_otchet:
                total_string.append(self.okonchanie("отчёт", total_otchet))
            else:
                total_string.append("")
            total_string.append("")
            if total_nanesenie:
                total_string.append(self.okonchanie("нанесение", total_nanesenie))
            else:
                total_string.append("")
            ws.append(total_string)
            ws_count_row.append(True)
            for _ in range(2):
                ws.append([" "])
                ws_count_row.append(False)
            chek_bold = True
            for index, row, row_numb in zip(ws_count_row, ws, itertools.count(1, 1)):
                if index:
                    for ind, cell in enumerate(row):
                        cell.border = self.thin_border
                        cell.alignment = Alignment(wrapText=True)
                        if chek_bold and ind != 0:
                            cell.font = Font(bold=True)
                    if chek_bold:
                        chek_bold = False
                else:
                    ws.merge_cells(f"A{row_numb}:F{row_numb}")
                    if not chek_bold:
                        chek_bold = True
        ws.page_setup.paperSize = "9"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 17
        ws.column_dimensions["F"].width = 18
        if personal_date:
            name = f"Отчёт с именами с {self.date_start.date()} по {self.date_end.date()}.xlsx"
        else:
            name = "I" * kvartal + f" квартал {year} года.xlsx"
        wb.save(name)
        return os.getcwd() + "\\" + name

    def make_excel_noname(self, kvartal, year, personal_date=False):
        """
        Функция, которая создает отчёт без имен.
        :param kvartal: Номер квартала (справочно, для подставления в название отчёта)
        :param year: Номер года (справочно, для подставления в название отчёта)
        :return: Путь до файла
        """
        ws_title = [
            "Темы",
            "Выпущенные опытно-промышленные партии",
            "Изготовленные лабораторные образцы и компоненты",
            "Написанные отчеты",
            "Выпущенные рецептуры и ТК",
            "Проведенные промышленные нанесения",
        ]
        wb = opx.Workbook()  # Создаем книгу эксель
        ws = wb.active  # запоминаем активный лист
        # добавляем в первую строку наименование таблицы
        if personal_date:
            ws.append(
                [
                    f"Таблица 1. Данные о результатах работы с {self.date_start.date()} по {self.date_end.date()}"
                ]
            )
        else:
            ws.append(
                [
                    f"Таблица 1. Данные о результатах работы за {kvartal} квартал {year} года"
                ]
            )
        ws.merge_cells(f"A1:F1")  # Объеди
        ws.append(ws_title)
        for ind in range(6):
            cell = ws[f"{chr(65 + ind)}2"]
            cell.font = Font(bold=True)
        ws_count_row = []
        for tema in self.thems:
            current_list_to_append = []
            if tema in self.text_of_tema_noname:
                current_list_to_append.insert(0, self.text_of_tema_noname[tema])
            else:
                current_list_to_append.insert(0, tema)
            cell_sintez = ""
            cell_obrazec = ""
            cell_otchet = ""
            cell_receptura = ""
            cell_nanesenie = ""
            sum_mass = 0
            sum_sintez = 0
            for name in self.thems[tema].workers:
                sintez_counter = SintezCounter()
                for i in self.thems[tema].workers[name].sinthesis:
                    sintez_counter.add_sintez(i)
                cell_sintez += str(sintez_counter)
                sum_mass += sintez_counter.summ_mass
                sum_sintez += sintez_counter.summ_sintes
                for i in self.short_show(
                    self.thems[tema].workers[name].plenki
                ) + self.short_show(self.thems[tema].workers[name].obrazci):
                    cell_obrazec += f"{i}, "
                for i in self.short_show_report(self.thems[tema].workers[name].otcheti):
                    cell_otchet += f"{i}, "
                for i in self.short_show(self.thems[tema].workers[name].naneseniya):
                    cell_nanesenie += f"{i}, "
            if cell_sintez:
                cell_sintez += (
                    f"\nИтого: "
                    + self.okonchanie("синтез", sum_sintez)
                    + f", {sum_mass} кг"
                )
            if cell_obrazec:
                cell_obrazec = cell_obrazec[:-2]
                cell_obrazec += f"\n\nИтого: " + self.okonchanie(
                    "образец", self.thems[tema].total_obraz
                )
            if cell_otchet:
                cell_otchet = cell_otchet[:-2]
                cell_otchet += f"\n\nИтого: " + self.okonchanie(
                    "отчёт", self.thems[tema].total_otchet
                )
            if cell_nanesenie:
                cell_nanesenie = cell_nanesenie[:-2]
                cell_nanesenie += f"\n\nИтого: " + self.okonchanie(
                    "нанесение", self.thems[tema].total_nanesenie
                )
            current_list_to_append.append(cell_sintez)
            current_list_to_append.append(cell_obrazec)
            current_list_to_append.append(cell_otchet)
            current_list_to_append.append(cell_receptura)
            current_list_to_append.append(cell_nanesenie)

            ws.append(current_list_to_append)
            ws_count_row.append(True)

            first_row = True
            for row in ws:
                if first_row:
                    first_row = False
                    continue
                for cell in row:
                    cell.border = self.thin_border
                    cell.alignment = Alignment(wrapText=True)

        ws.page_setup.paperSize = "9"

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 17
        ws.column_dimensions["F"].width = 18
        if personal_date:
            name = f"Отчёт без имен с {self.date_start.date()} по {self.date_end.date()}.xlsx"
        else:
            name = f"О работе за {kvartal} квартал {year} года.xlsx"
        wb.save(name)
        return os.getcwd() + "\\" + name


class Tema:
    """
    Класс тема. Содержит работников, у которых внутри их работы по этой теме.
    """

    def __init__(self, tema_name):
        self.tema_name = tema_name
        self.old_name = None
        self.workers = {}
        self.text_with_names = ""
        self.text_no_names = ""
        self.total_obraz = 0
        self.total_otchet = 0
        self.total_nanesenie = 0
        self.total_sintes_mass = 0
        self.total_sintes_count = 0

    def add_item(self, item: (Shablon, Sintez)):
        """
        Принимает в тему работу и отдает её соответствующему работнику
        :param item: работа
        :return: None
        """
        if item.author not in self.workers:
            self.workers[item.author] = Worker(item.author)

        if type(item) == Plenka:
            self.workers[item.author].plenki.append(item)
            self.total_obraz += 1

        elif type(item) == Otchet:
            self.workers[item.author].otcheti.append(item)
            self.total_otchet += 1

        elif type(item) == Obrazec:
            self.workers[item.author].obrazci.append(item)
            self.total_obraz += 1

        elif type(item) == Nanesenie:
            self.workers[item.author].naneseniya.append(item)
            self.total_nanesenie += 1

        elif type(item) == Sintez:
            self.workers[item.author].sinthesis.append(item)
            self.total_sintes_count += 1
            self.total_sintes_mass += item.mass
            self.total_sintes_mass = self.total_sintes_mass * 1000000 // 1000 / 1000


class SintezCounter:
    """
    Класс-обработчик для подсчёта синтезов и вывода соответствующей строки для дальнейшей вставки в отчёт
    """

    def __init__(self):
        self.tema = {}
        self.summ_mass = 0
        self.summ_sintes = 0

    def add_sintez(self, sintez: Sintez):
        """
        Добавляет синтез в обработчик для учёта
        :param sintez: объект класса Sintez
        :return: None
        """
        if sintez.base_tema not in self.tema:
            self.tema[sintez.base_tema] = {}
        if sintez.komponent not in self.tema[sintez.base_tema]:
            self.tema[sintez.base_tema][sintez.komponent] = {
                "mass": sintez.mass,
                "count": 1,
            }
        else:
            self.tema[sintez.base_tema][sintez.komponent]["mass"] += sintez.mass
            self.tema[sintez.base_tema][sintez.komponent]["count"] += 1

        self.summ_mass += sintez.mass
        self.summ_sintes += 1

    def __str__(self):
        str_to_return = ""
        for tema in self.tema:
            str_to_return += tema + "\n\t"
            for komponent in self.tema[tema]:
                str_to_return += (
                    komponent
                    + f"\n"
                    + Cunductor.okonchanie(
                        "синтез", self.tema[tema][komponent]["count"]
                    )
                    + ", "
                    + str(self.tema[tema][komponent]["mass"])
                    + " кг\n\t"
                )
        return str_to_return

    def str_one_tema(self):
        str_to_return = ""
        for tema in self.tema:
            for komponent in self.tema[tema]:
                str_to_return += (
                    komponent
                    + f"\n  "
                    + Cunductor.okonchanie(
                        "синтез", self.tema[tema][komponent]["count"]
                    )
                    + ", "
                    + str(self.tema[tema][komponent]["mass"])
                    + " кг\n"
                )
        return str_to_return

    def return_all(self):
        return self.summ_mass, self.summ_sintes
